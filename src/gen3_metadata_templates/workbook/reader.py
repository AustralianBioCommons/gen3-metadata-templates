"""Read a filled-in workbook back into per-node records, remembering where every
value came from.

The critical output is the coordinate map: for each (node, record index,
property) the reader records the exact cell (sheet + A1 reference). That is what
lets validation translate an engine error like "'ten' is not of type 'integer'"
into "Sheet subject, cell D5". Values are coerced toward their schema types only
where unambiguous; anything that can't be coerced is passed through untouched so
the validator reports it rather than the reader masking it.
"""

from __future__ import annotations

import datetime as _dt
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import openpyxl

from gen3_metadata_templates.constants import (
    FIRST_DATA_ROW,
    HINT_ROW,
    LIST_SPLIT_CHAR,
    META_SHEET,
)
from gen3_metadata_templates.model import ColumnKind, ColumnSpec, NodeTemplate, TemplateSpec


@dataclass(frozen=True)
class CellRef:
    """A single cell location, for pointing a user at exactly where to fix."""

    sheet: str
    row: int  # 1-indexed Excel row
    column_letter: str

    @property
    def a1(self) -> str:
        return f"{self.column_letter}{self.row}"


@dataclass
class ParsedWorkbook:
    """Records extracted from a workbook, plus a coordinate map and warnings."""

    records: Dict[str, List[dict]] = field(default_factory=dict)
    warnings: List[str] = field(default_factory=list)
    # node -> list (per record) of {prop_name: CellRef}
    _coords: Dict[str, List[Dict[str, CellRef]]] = field(default_factory=dict)
    # node -> {prop_name: header} for columns that were expected but missing
    missing_columns: Dict[str, List[str]] = field(default_factory=dict)

    def coord(self, node: str, index: int, prop: str) -> Optional[CellRef]:
        """The cell a value came from, or None if it can't be located."""
        node_coords = self._coords.get(node)
        if node_coords is None or index >= len(node_coords):
            return None
        return node_coords[index].get(prop)


def _col_letter(index_zero_based: int) -> str:
    return openpyxl.utils.get_column_letter(index_zero_based + 1)


def coerce_cell(value: Any, col: ColumnSpec) -> Any:
    """Convert a raw cell value toward the column's schema type.

    Coercion is deliberately conservative: it fixes the mismatches Excel
    introduces (numbers arriving as floats, dates as datetimes) but never forces
    a value that doesn't fit — a non-numeric string in an integer column is
    returned unchanged so the validator flags it.
    """
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None

    if col.kind is ColumnKind.LINK:
        return _fold_link(value, col)

    if col.is_multi:  # array-valued property
        return _split_list(value, col)

    return _coerce_scalar(value, col.data_type)


def _coerce_scalar(value: Any, data_type: str) -> Any:
    if data_type == "integer":
        if isinstance(value, bool):
            return value
        if isinstance(value, float) and value.is_integer():
            return int(value)
        if isinstance(value, str):
            try:
                return int(value)
            except ValueError:
                return value
        return value

    if data_type == "number":
        if isinstance(value, str):
            try:
                return float(value)
            except ValueError:
                return value
        return value

    if data_type == "boolean":
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)) and value in (0, 1):
            return bool(value)
        if isinstance(value, str):
            lowered = value.lower()
            if lowered in ("true", "yes"):
                return True
            if lowered in ("false", "no"):
                return False
        return value

    # string / enum: normalise numbers and dates to clean text.
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return str(int(value)) if float(value).is_integer() else str(value)
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat()
    return value


def _split_list(value: Any, col: ColumnSpec) -> Any:
    """Split a ';'-separated cell into a list, coercing each item."""
    if not isinstance(value, str):
        return _coerce_scalar(value, "string")
    parts = [p.strip() for p in value.split(LIST_SPLIT_CHAR) if p.strip()]
    item_type = "string"  # array item type; enums/strings both read as text here
    return [_coerce_scalar(p, item_type) for p in parts]


def _fold_link(value: Any, col: ColumnSpec) -> Any:
    """Turn a foreign-key cell into the reference shape gen3_validator expects.

    A single parent becomes ``{"submitter_id": "id"}``; a to-many link with a
    ';'-separated cell becomes a list of such dicts.
    """
    text = str(int(value)) if isinstance(value, float) and value.is_integer() else str(value)
    if col.is_multi:
        parts = [p.strip() for p in text.split(LIST_SPLIT_CHAR) if p.strip()]
        return [{"submitter_id": p} for p in parts]
    return {"submitter_id": text}


def read_meta(workbook_path: Union[str, Path]) -> Optional[dict]:
    """Read the hidden ``_g3mt`` metadata sheet, if present.

    Returns a dict with ``schema_file``/``target_node``/``path`` and a
    ``node_sheets`` map, or None if the workbook has no metadata sheet.
    """
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        if META_SHEET not in wb.sheetnames:
            return None
        ws = wb[META_SHEET]
        meta: Dict[str, Any] = {}
        node_sheets: Dict[str, str] = {}
        in_map = False
        for row in ws.iter_rows(values_only=True):
            key, value = (row + (None, None))[:2]
            if key is None:
                continue
            if key == "node" and value == "sheet":
                in_map = True
                continue
            if in_map:
                node_sheets[str(key)] = str(value)
            else:
                meta[str(key)] = value
        meta["node_sheets"] = node_sheets
        return meta
    finally:
        wb.close()


def read_workbook(workbook_path: Union[str, Path], spec: TemplateSpec) -> ParsedWorkbook:
    """Parse a filled workbook into per-node records using ``spec`` as the map."""
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    parsed = ParsedWorkbook()
    try:
        for node_template in spec.nodes:
            _read_node_sheet(wb, node_template, parsed)
    finally:
        wb.close()
    return parsed


def _read_node_sheet(wb, node_template: NodeTemplate, parsed: ParsedWorkbook) -> None:
    node = node_template.node
    sheet_name = node_template.sheet_name
    if sheet_name not in wb.sheetnames:
        parsed.warnings.append(
            f"Sheet '{sheet_name}' (node '{node}') is missing — its records were skipped."
        )
        parsed.records[node] = []
        parsed._coords[node] = []
        return

    ws = wb[sheet_name]

    # Map header text -> (column index, ColumnSpec) from row 1.
    header_to_col: Dict[int, ColumnSpec] = {}
    seen_headers = set()
    for col_idx in range(ws.max_column):
        header = ws.cell(HINT_ROW - 1, col_idx + 1).value  # HEADER_ROW == 1
        if header is None:
            continue
        header = str(header).strip()
        spec_col = node_template.column_by_header(header)
        if spec_col is None:
            parsed.warnings.append(
                f"Column '{header}' on sheet '{sheet_name}' is not in the schema — ignored."
            )
            continue
        header_to_col[col_idx] = spec_col
        seen_headers.add(header)

    missing = [c.prop_name for c in node_template.columns if c.header not in seen_headers]
    if missing:
        parsed.missing_columns[node] = missing

    records: List[dict] = []
    coords: List[Dict[str, CellRef]] = []
    for row_idx in range(FIRST_DATA_ROW, ws.max_row + 1):
        record, row_coords = _read_row(ws, row_idx, header_to_col, sheet_name)
        if record is None:
            continue
        record["type"] = node
        records.append(record)
        coords.append(row_coords)

    parsed.records[node] = records
    parsed._coords[node] = coords


def _read_row(ws, row_idx: int, header_to_col: Dict[int, ColumnSpec], sheet_name: str):
    """Read one data row; return (record, coord_map) or (None, {}) if blank."""
    record: dict = {}
    row_coords: Dict[str, CellRef] = {}
    any_value = False

    for col_idx, spec_col in header_to_col.items():
        raw = ws.cell(row_idx, col_idx + 1).value
        ref = CellRef(sheet_name, row_idx, _col_letter(col_idx))
        # Record the coordinate even for blanks so "required but empty" errors
        # can still point at the exact cell.
        row_coords[spec_col.prop_name] = ref

        coerced = coerce_cell(raw, spec_col)
        if coerced is None or coerced == []:
            continue
        any_value = True
        record[spec_col.prop_name] = coerced

    if not any_value:
        return None, {}
    return record, row_coords
