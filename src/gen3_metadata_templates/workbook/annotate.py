"""Write a copy of a validated workbook with the problem cells highlighted.

The annotated file is a review aid: each flagged cell is filled red and carries
a comment explaining the problem, and a summary sheet lists every finding with a
link to its cell. The user keeps filling in their original file; this copy just
shows them what to fix.
"""

from __future__ import annotations

from pathlib import Path
from typing import Union

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill

from gen3_metadata_templates.errors import G3mtError
from gen3_metadata_templates.validation.report import ValidationReport

_BAD_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_SUMMARY_SHEET = "Validation Errors"


def write_annotated_copy(
    workbook_path: Union[str, Path],
    report: ValidationReport,
    output_path: Union[str, Path],
) -> None:
    """Write ``workbook_path`` to ``output_path`` with findings marked up.

    :raises G3mtError: if the output path is the same as the input (which would
        destroy the user's original file).
    """
    if Path(workbook_path).resolve() == Path(output_path).resolve():
        raise G3mtError("Annotated copy must be written to a different file than the input.")

    wb = openpyxl.load_workbook(workbook_path)

    for finding in report.findings:
        if finding.cell is None or finding.sheet not in wb.sheetnames:
            continue
        ws = wb[finding.sheet]
        cell = ws[finding.cell.a1]
        cell.fill = _BAD_FILL
        cell.comment = Comment(finding.message, "g3mt")

    _write_summary(wb, report)
    wb.save(output_path)


def _write_summary(wb, report: ValidationReport) -> None:
    if _SUMMARY_SHEET in wb.sheetnames:
        del wb[_SUMMARY_SHEET]
    ws = wb.create_sheet(_SUMMARY_SHEET, 0)

    bold = Font(bold=True)
    for col, title in enumerate(["Location", "Column", "Problem"], start=1):
        cell = ws.cell(1, col, title)
        cell.font = bold
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 80

    for row, finding in enumerate(report.findings, start=2):
        if finding.cell is not None:
            link = f"#'{finding.sheet}'!{finding.cell.a1}"
            location = ws.cell(row, 1, finding.location)
            location.hyperlink = link
            location.font = Font(color="0563C1", underline="single")
        else:
            ws.cell(row, 1, finding.location)
        ws.cell(row, 2, finding.header or "-")
        ws.cell(row, 3, finding.message)
