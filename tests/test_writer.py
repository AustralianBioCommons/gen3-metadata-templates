"""Tests for :mod:`gen3_metadata_templates.workbook.writer`.

The writer's output is what a submitter actually opens, so these tests generate
a real workbook and re-open it with openpyxl to confirm the structure: sheet
order, headers and hint rows, the guide/metadata sheets, cross-sheet link
dropdowns, and enum dropdowns. Getting the dropdowns right is the feature that
makes linked-data submission approachable, so they are checked explicitly.
"""

from __future__ import annotations

import openpyxl
import pytest

from gen3_metadata_templates import (
    build_multi_template_spec,
    build_template_spec,
    write_template,
)
from gen3_metadata_templates.constants import (
    DEFAULT_EXCLUDED_NODES,
    DICTIONARY_SHEET,
    INSTRUCTIONS_SHEET,
    META_SHEET,
)
from gen3_metadata_templates.selection import resolve_selection


@pytest.fixture()
def sample_workbook(mini_bundle, tmp_path):
    """Generate a template for the subject->visit->sample path and return its path."""
    spec = build_template_spec(mini_bundle, "sample", ["subject", "visit", "sample"])
    out = tmp_path / "sample_template.xlsx"
    write_template(spec, out, data_rows=50)
    return out, spec


def test_sheets_are_in_path_order_with_guides(sample_workbook):
    """Node sheets appear in fill order, wrapped by the guide sheets.

    Instructions must come first (it's the first thing a user should read) and
    the node sheets must be in path order so a submitter naturally fills parents
    before children.
    """
    path, _ = sample_workbook
    wb = openpyxl.load_workbook(path)
    names = wb.sheetnames
    assert names[0] == INSTRUCTIONS_SHEET
    assert names.index("subject") < names.index("visit") < names.index("sample")
    assert DICTIONARY_SHEET in names
    assert META_SHEET in names


def test_header_and_hint_rows(sample_workbook):
    """Row 1 holds headers; row 2 holds the type/requirement hint."""
    path, _ = sample_workbook
    wb = openpyxl.load_workbook(path)
    ws = wb["sample"]
    assert ws.cell(1, 1).value == "submitter_id"
    assert ws.cell(1, 2).value == "subject.submitter_id"
    assert "required" in ws.cell(2, 1).value


def test_foreign_key_dropdown_targets_parent_sheet(sample_workbook):
    """The FK column's dropdown pulls IDs from the parent sheet's column.

    This cross-sheet dropdown is what lets a submitter pick a real parent
    submitter_id instead of typing (and mistyping) it, so it must reference the
    parent's named range.
    """
    path, _ = sample_workbook
    wb = openpyxl.load_workbook(path)
    ws = wb["sample"]
    sources = {dv.formula1 for dv in ws.data_validations.dataValidation}
    assert "ids_subject" in sources
    assert "ids_visit" in sources
    # And the defined name actually points at the subject sheet's column A.
    assert "subject" in str(wb.defined_names["ids_subject"].value)


def test_enum_dropdown_lists_allowed_values(sample_workbook):
    """An enum column offers exactly its allowed values as a dropdown."""
    path, _ = sample_workbook
    wb = openpyxl.load_workbook(path)
    ws = wb["sample"]
    enum_formulas = [
        dv.formula1 for dv in ws.data_validations.dataValidation if "Blood" in str(dv.formula1)
    ]
    assert enum_formulas, "expected a dropdown containing the sample_type values"
    assert "Tissue" in enum_formulas[0] and "Saliva" in enum_formulas[0]


def test_dictionary_has_one_row_per_column(sample_workbook):
    """The Dictionary sheet documents every column across every node sheet.

    It's the submitter's reference for what each field means, so it must have a
    row for each ColumnSpec (plus the header row).
    """
    path, spec = sample_workbook
    wb = openpyxl.load_workbook(path)
    ws = wb[DICTIONARY_SHEET]
    total_columns = sum(len(nt.columns) for nt in spec.nodes)
    assert ws.max_row == total_columns + 1  # + header row


def test_meta_sheet_records_target_path_and_node_order(sample_workbook):
    """The hidden metadata sheet lets validate recover the sheets a workbook holds.

    ``node_order`` is the authoritative list. ``target_node`` and ``path`` are
    kept alongside it so that an older g3mt install can still read a workbook
    written by this one — deleting them would break that.
    """
    path, _ = sample_workbook
    wb = openpyxl.load_workbook(path)
    ws = wb[META_SHEET]
    meta = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    assert meta["target_node"] == "sample"
    assert meta["path"] == "subject,visit,sample"
    assert meta["node_order"] == "subject,visit,sample"
    assert meta["target_nodes"] == "sample"
    assert str(meta["meta_format"]) == "2"


def test_meta_sheet_target_paths_round_trip_as_json(sample_workbook):
    """The per-target paths survive as parseable JSON.

    Node names can contain almost anything, so the per-target paths are stored
    as JSON rather than a delimited string that could be ambiguous to split.
    """
    import json

    path, _ = sample_workbook
    wb = openpyxl.load_workbook(path)
    ws = wb[META_SHEET]
    meta = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    assert json.loads(meta["target_paths"]) == {"sample": ["subject", "visit", "sample"]}


def test_meta_sheet_records_schema_version_and_source(sample_workbook):
    """The workbook records which schema version and source it was generated from.

    This is what lets a person (and the validator) tell which dictionary version
    a filled template corresponds to, and where the schema came from.
    """
    path, spec = sample_workbook
    wb = openpyxl.load_workbook(path)
    ws = wb[META_SHEET]
    meta = {ws.cell(r, 1).value: ws.cell(r, 2).value for r in range(1, ws.max_row + 1)}
    assert meta["schema_version"] == "0.1.0"
    assert meta["schema_source"] == spec.schema_path


def test_instructions_sheet_shows_schema_version(sample_workbook):
    """A person opening the workbook can see the schema version on the Instructions sheet."""
    path, _ = sample_workbook
    wb = openpyxl.load_workbook(path)
    ws = wb["Instructions"]
    text = "\n".join(
        str(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value
    )
    assert "0.1.0" in text


# --- multi-node templates -------------------------------------------------


@pytest.fixture()
def clinical_workbook(clinical_hub_bundle, tmp_path):
    """A multi-node template covering a whole clinical category."""
    selection = resolve_selection(
        clinical_hub_bundle,
        clinical_hub_bundle.nodes_in_category("clinical"),
        excluded_nodes=DEFAULT_EXCLUDED_NODES,
        category="clinical",
    )
    spec = build_multi_template_spec(clinical_hub_bundle, selection)
    out = tmp_path / "clinical_template.xlsx"
    write_template(spec, out, data_rows=20)
    return out, spec


def test_sheet_order_matches_the_resolved_order(clinical_workbook):
    """Sheets appear in the order the selection resolved, parents first."""
    path, spec = clinical_workbook
    names = openpyxl.load_workbook(path).sheetnames
    node_positions = [names.index(nt.sheet_name) for nt in spec.nodes]
    assert node_positions == sorted(node_positions)
    assert names.index("subject") < names.index("clinical_descriptor")
    assert names.index("clinical_descriptor") < names.index("demographic")


def test_instructions_show_an_indented_fill_order(clinical_workbook):
    """A branching template is drawn as an indented tree, not a single line.

    With several sheets hanging off one parent, `a -> b -> c` would be a lie.
    Real Excel indentation shows which sheets are siblings and which are nested,
    so a submitter can see at a glance what depends on what.
    """
    path, _ = clinical_workbook
    ws = openpyxl.load_workbook(path)["Instructions"]
    indents = {}
    for row in range(1, ws.max_row + 1):
        cell = ws.cell(row, 1)
        if cell.value in ("subject", "clinical_descriptor", "demographic"):
            indents[cell.value] = int(cell.alignment.indent or 0)

    assert indents["subject"] < indents["clinical_descriptor"] < indents["demographic"]


def test_instructions_say_you_need_not_fill_every_sheet(clinical_workbook):
    """A whole-category template must say that unused sheets can be left empty.

    Asking for every clinical node gives sheets a given study may have no data
    for. Without this line a researcher may think the submission is incomplete.
    """
    path, _ = clinical_workbook
    ws = openpyxl.load_workbook(path)["Instructions"]
    text = " ".join(
        str(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value
    )
    assert "do not have to fill in every sheet" in text


def test_single_target_instructions_keep_the_simple_one_line_order(sample_workbook):
    """A single-target template keeps its original, simpler wording.

    Most templates cover one straight chain, and an indented tree would be more
    ceremony than that needs.
    """
    path, _ = sample_workbook
    ws = openpyxl.load_workbook(path)["Instructions"]
    text = " ".join(
        str(ws.cell(r, 1).value) for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value
    )
    assert "subject -> visit -> sample" in text
    assert "do not have to fill in every sheet" not in text


def test_a_node_named_like_a_guide_sheet_does_not_collide():
    """A node called 'Dictionary' must not overwrite the real Dictionary sheet.

    Sheet names are assigned from node names, and the workbook reserves four of
    its own. A collision would silently destroy a guide sheet.
    """
    from gen3_metadata_templates.workbook.naming import sheet_names

    mapping = sheet_names(["Dictionary", "Instructions", "subject"])
    assert mapping["Dictionary"] != "Dictionary"
    assert mapping["Instructions"] != "Instructions"
    assert mapping["subject"] == "subject"
