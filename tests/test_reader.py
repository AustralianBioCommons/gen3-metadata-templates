"""Tests for :mod:`gen3_metadata_templates.workbook.reader`.

The reader's two jobs are (1) coerce raw Excel values toward their schema types
without ever masking a genuine mistake, and (2) remember the exact cell each
value came from so validation can point a user at it. Coercion is tested as a
pure function; the coordinate/record building is tested against a real
generated-then-filled workbook.
"""

from __future__ import annotations

import datetime

import openpyxl
import pytest

from gen3_metadata_templates import build_template_spec, write_template
from gen3_metadata_templates.model import ColumnKind, ColumnSpec
from gen3_metadata_templates.workbook.reader import coerce_cell, read_meta, read_workbook


def _col(data_type="string", kind=ColumnKind.PROPERTY, is_multi=False):
    return ColumnSpec(
        header="h",
        prop_name="p",
        kind=kind,
        data_type=data_type,
        required=False,
        is_multi=is_multi,
    )


# --- coerce_cell: the pure coercion rules ---------------------------------


def test_integer_float_becomes_int():
    """Excel stores whole numbers as floats; an integer column should read int.

    Without this, ``42`` would arrive as ``42.0`` and fail integer validation
    even though the user typed a valid whole number.
    """
    assert coerce_cell(42.0, _col("integer")) == 42
    assert isinstance(coerce_cell(42.0, _col("integer")), int)


def test_uncoercible_integer_passes_through_unchanged():
    """A non-numeric value in an integer column is left as-is.

    The reader must not hide the error — passing ``"ten"`` straight through lets
    the validator report a clear type error at that cell.
    """
    assert coerce_cell("ten", _col("integer")) == "ten"


def test_boolean_words_become_bool():
    """Human spellings of true/false map to real booleans."""
    assert coerce_cell("TRUE", _col("boolean")) is True
    assert coerce_cell("no", _col("boolean")) is False


def test_number_in_string_column_is_clean_text():
    """A number typed in a string column becomes tidy text, not '2023.0'."""
    assert coerce_cell(2023.0, _col("string")) == "2023"


def test_datetime_becomes_iso_string():
    """Date cells are serialised to ISO 8601 strings for schema validation."""
    value = datetime.datetime(2024, 1, 31, 9, 0, 0)
    assert coerce_cell(value, _col("string")) == "2024-01-31T09:00:00"


def test_blank_cell_is_none():
    """Whitespace-only and empty cells read as None (i.e. not provided)."""
    assert coerce_cell("   ", _col("string")) is None
    assert coerce_cell(None, _col("string")) is None


def test_array_cell_splits_on_separator():
    """A ';'-separated cell becomes a list of trimmed values."""
    col = _col("array", is_multi=True)
    assert coerce_cell("a; b ;c", col) == ["a", "b", "c"]


def test_link_folds_to_submitter_id_dict():
    """A single foreign key folds to the reference shape the validator expects."""
    col = _col("string", kind=ColumnKind.LINK)
    assert coerce_cell("subj_1", col) == {"submitter_id": "subj_1"}


def test_to_many_link_folds_to_list():
    """A to-many foreign key with a ';'-separated cell folds to a list of refs."""
    col = _col("string", kind=ColumnKind.LINK, is_multi=True)
    assert coerce_cell("a; b", col) == [{"submitter_id": "a"}, {"submitter_id": "b"}]


# --- read_workbook: records + coordinates on a real workbook --------------


@pytest.fixture()
def filled_workbook(mini_bundle, tmp_path):
    """Generate a template, fill one subject and one linked sample, return path+spec."""
    spec = build_template_spec(mini_bundle, "sample", ["subject", "visit", "sample"])
    out = tmp_path / "filled.xlsx"
    write_template(spec, out, data_rows=20)

    wb = openpyxl.load_workbook(out)
    subj = wb["subject"]
    subj_hdr = {subj.cell(1, c).value: c for c in range(1, subj.max_column + 1)}
    subj.cell(3, subj_hdr["submitter_id"]).value = "subj_1"
    subj.cell(3, subj_hdr["subject_id"]).value = "S1"
    subj.cell(3, subj_hdr["age"]).value = 42

    samp = wb["sample"]
    samp_hdr = {samp.cell(1, c).value: c for c in range(1, samp.max_column + 1)}
    samp.cell(3, samp_hdr["submitter_id"]).value = "samp_1"
    samp.cell(3, samp_hdr["subject.submitter_id"]).value = "subj_1"
    samp.cell(3, samp_hdr["sample_id"]).value = "SAMP1"
    samp.cell(3, samp_hdr["sample_type"]).value = "Blood"
    wb.save(out)
    return out, spec


def test_records_carry_type_and_folded_links(filled_workbook):
    """Each record gets its node ``type`` and folded link references.

    ``validate_list_dict`` requires the ``type`` field, and link validation
    requires the ``{"submitter_id": ...}`` shape, so both must be present after
    reading.
    """
    path, spec = filled_workbook
    parsed = read_workbook(path, spec)
    sample = parsed.records["sample"][0]
    assert sample["type"] == "sample"
    assert sample["subjects"] == {"submitter_id": "subj_1"}


def test_blank_rows_are_skipped(filled_workbook):
    """Only rows with data become records; the empty provisioned rows don't."""
    path, spec = filled_workbook
    parsed = read_workbook(path, spec)
    assert len(parsed.records["subject"]) == 1
    assert len(parsed.records["sample"]) == 1


def test_coordinates_point_at_the_source_cell(filled_workbook):
    """The coordinate map returns the exact A1 cell a value was read from.

    This is the mechanism that turns a validation error into "fix cell C3", so
    it must resolve to the right cell.
    """
    path, spec = filled_workbook
    parsed = read_workbook(path, spec)
    ref = parsed.coord("subject", 0, "age")
    assert ref is not None
    assert ref.sheet == "subject"
    assert ref.row == 3


def test_read_meta_recovers_target_and_path(filled_workbook):
    """The metadata sheet is read back into a usable dict for validate."""
    path, _ = filled_workbook
    meta = read_meta(path)
    assert meta["target_node"] == "sample"
    assert meta["node_sheets"]["subject"] == "subject"
