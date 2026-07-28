"""Integration smoke tests against the real 34-node ACDC schema.

The mini schema is deliberately small; these tests guard against regressions on
a full, real-world Gen3 dictionary — deep paths, many nodes, real enums — that
the hand-built fixture can't represent. They intentionally assert only broad,
stable properties so they don't turn brittle as the example schema evolves.
"""

from __future__ import annotations

import openpyxl

from gen3_metadata_templates import (
    SchemaBundle,
    build_multi_template_spec,
    build_template_spec,
    enumerate_paths,
    resolve_path,
    validate_workbook,
    write_template,
)
from gen3_metadata_templates.constants import DEFAULT_EXCLUDED_NODES
from gen3_metadata_templates.selection import resolve_selection


def test_generate_deep_node_writes_expected_sheets(acdc_schema_path, tmp_path):
    """A deep data-file node should produce a multi-sheet workbook in path order.

    ``lipidomics_file`` sits several links below the root, so its template should
    contain a sheet for each ancestor on the chosen path plus the guide sheets —
    proving generation scales beyond the toy fixture.
    """
    bundle = SchemaBundle(acdc_schema_path)
    paths = enumerate_paths(bundle, "lipidomics_file", DEFAULT_EXCLUDED_NODES)
    chosen = resolve_path(paths, path_arg="1")
    spec = build_template_spec(bundle, "lipidomics_file", chosen)

    out = tmp_path / "lipidomics_file.xlsx"
    write_template(spec, out, data_rows=50)

    names = openpyxl.load_workbook(out).sheetnames
    assert "lipidomics_file" in names
    for node in chosen:
        if node not in DEFAULT_EXCLUDED_NODES:
            assert node in names


def test_empty_generated_template_validates_clean(acdc_schema_path, tmp_path):
    """A freshly generated (unfilled) template must validate with no errors.

    With no data rows there is nothing to violate the schema, so validation
    should pass — a good end-to-end check that generate and validate agree on
    the workbook format for a real schema.
    """
    bundle = SchemaBundle(acdc_schema_path)
    chosen = resolve_path(
        enumerate_paths(bundle, "demographic", DEFAULT_EXCLUDED_NODES), path_arg="1"
    )
    spec = build_template_spec(bundle, "demographic", chosen)
    out = tmp_path / "demographic.xlsx"
    write_template(spec, out, data_rows=20)

    report = validate_workbook(out, acdc_schema_path)
    assert report.ok, [f.message for f in report.findings]


# --- whole-category templates on the real dictionary ----------------------
#
# Expectations here are derived from the schema itself (via the category API),
# never from a hardcoded node list. Gen3 dictionaries get restructured — this
# example schema uses a `timepoint` hub where newer ones use
# `clinical_descriptor` — and a test that pins node names would break on a
# schema refresh while telling us nothing about the code.


def test_clinical_category_produces_a_sheet_for_every_clinical_node(acdc_bundle, tmp_path):
    """Selecting the clinical category gives one sheet per clinical node.

    This is the researcher's headline command on a real dictionary, so it has to
    hold whatever shape that dictionary happens to have.
    """
    clinical = acdc_bundle.nodes_in_category("clinical")
    selection = resolve_selection(
        acdc_bundle, clinical, excluded_nodes=DEFAULT_EXCLUDED_NODES, category="clinical"
    )
    spec = build_multi_template_spec(acdc_bundle, selection)
    out = tmp_path / "clinical.xlsx"
    write_template(spec, out, data_rows=20)

    names = openpyxl.load_workbook(out).sheetnames
    for node in clinical:
        assert node in names, f"expected a sheet for clinical node {node}"


def test_every_parent_comes_before_its_child(acdc_bundle):
    """The sheet order must satisfy the one invariant that actually matters.

    Rather than pinning specific node names, assert the property: for every
    node in the template, each of its parents that is also in the template
    appears earlier. This holds on any dictionary shape, so it catches a broken
    ordering without breaking when the schema is restructured.
    """
    selection = resolve_selection(
        acdc_bundle,
        acdc_bundle.nodes_in_category("clinical"),
        excluded_nodes=DEFAULT_EXCLUDED_NODES,
    )
    position = {node: i for i, node in enumerate(selection.nodes)}
    for node in selection.nodes:
        for link in acdc_bundle.links(node):
            if link.target_type in position:
                assert position[link.target_type] < position[node], (
                    f"{link.target_type} must come before its child {node}"
                )


def test_clinical_category_template_validates_clean(acdc_bundle, tmp_path):
    """A freshly generated whole-category template has nothing wrong with it.

    With no data rows there is nothing to violate the schema, so this proves
    generate and validate agree about a large multi-node workbook.
    """
    selection = resolve_selection(
        acdc_bundle,
        acdc_bundle.nodes_in_category("clinical"),
        excluded_nodes=DEFAULT_EXCLUDED_NODES,
        category="clinical",
    )
    spec = build_multi_template_spec(acdc_bundle, selection)
    out = tmp_path / "clinical.xlsx"
    write_template(spec, out, data_rows=20)

    report = validate_workbook(out, str(acdc_bundle.schema_path))
    assert report.ok, [f.message for f in report.findings]


def test_subject_alone_generates_now_that_roots_are_supported(acdc_bundle, tmp_path):
    """A root node can have a template of its own.

    ``subject`` has no parent once the administrative nodes are excluded. This
    used to fail outright, which also made any category containing a root node
    impossible to select.
    """
    selection = resolve_selection(acdc_bundle, ["subject"], excluded_nodes=DEFAULT_EXCLUDED_NODES)
    spec = build_multi_template_spec(acdc_bundle, selection)
    out = tmp_path / "subject.xlsx"
    write_template(spec, out, data_rows=10)
    assert "subject" in openpyxl.load_workbook(out).sheetnames


def test_data_file_category_generates_and_validates_clean(acdc_bundle, tmp_path):
    """The data_file category stresses excluded parents and required subgroups.

    Many data-file nodes have a required subgroup link to
    ``core_metadata_collection``, which is excluded by default — so its column is
    correctly absent. Validation must not then demand it, or every such workbook
    would be unfixable.
    """
    selection = resolve_selection(
        acdc_bundle,
        acdc_bundle.nodes_in_category("data_file"),
        excluded_nodes=DEFAULT_EXCLUDED_NODES,
        category="data_file",
    )
    spec = build_multi_template_spec(acdc_bundle, selection)
    out = tmp_path / "data_files.xlsx"
    write_template(spec, out, data_rows=10)

    report = validate_workbook(out, str(acdc_bundle.schema_path))
    assert report.ok, [f.message for f in report.findings]
