"""Workbooks made by older versions of g3mt must keep validating.

g3mt 2.2.0 and earlier recorded only a single ``target_node`` and ``path`` in the
hidden ``_g3mt`` sheet. People have those files on disk, half filled in. This
module is the contract that they keep working exactly as they did — not merely
"without crashing", but rebuilding the *same* column layout, so an error still
lands on the same cell.

``_downgrade_meta_to_2_2_0`` rewrites a freshly generated workbook's metadata to
the old shape, which is how we test against the old format without keeping a
binary fixture around.
"""

from __future__ import annotations

import openpyxl
import pytest

from gen3_metadata_templates import build_template_spec, validate_workbook, write_template
from gen3_metadata_templates.constants import META_SHEET
from gen3_metadata_templates.errors import WorkbookFormatError

LEGACY_KEYS = ("g3mt_version", "schema_file", "target_node", "path", "data_rows")


def _read_meta_rows(path):
    """Return the ``_g3mt`` key/value rows as a dict (ignoring the node map)."""
    wb = openpyxl.load_workbook(path)
    ws = wb[META_SHEET]
    rows = {}
    for r in range(1, ws.max_row + 1):
        key, value = ws.cell(r, 1).value, ws.cell(r, 2).value
        if key == "node" and value == "sheet":
            break
        if key:
            rows[str(key)] = value
    wb.close()
    return rows


def _downgrade_meta_to_2_2_0(path):
    """Rewrite a workbook's metadata sheet into the g3mt 2.2.0 shape.

    Keeps only the keys 2.2.0 wrote (dropping node_order, target_nodes and the
    rest) plus the node -> sheet map, so the file is indistinguishable from one
    produced by that release.
    """
    wb = openpyxl.load_workbook(path)
    ws = wb[META_SHEET]

    existing = {}
    node_sheets = []
    in_map = False
    for r in range(1, ws.max_row + 1):
        key, value = ws.cell(r, 1).value, ws.cell(r, 2).value
        if key == "node" and value == "sheet":
            in_map = True
            continue
        if key is None:
            continue
        if in_map:
            node_sheets.append((key, value))
        else:
            existing[str(key)] = value

    wb.remove(ws)
    ws = wb.create_sheet(META_SHEET)
    row = 1
    for key in LEGACY_KEYS:
        ws.cell(row, 1, key)
        ws.cell(row, 2, existing.get(key, ""))
        row += 1
    row += 1  # blank separator, as the writer emits
    ws.cell(row, 1, "node")
    ws.cell(row, 2, "sheet")
    for node, sheet in node_sheets:
        row += 1
        ws.cell(row, 1, node)
        ws.cell(row, 2, sheet)
    ws.sheet_state = "hidden"
    wb.save(path)


@pytest.fixture()
def legacy_workbook(mini_bundle, tmp_path):
    """A workbook whose metadata looks exactly like g3mt 2.2.0 wrote it."""
    spec = build_template_spec(mini_bundle, "sample", ["subject", "visit", "sample"])
    out = tmp_path / "legacy.xlsx"
    write_template(spec, out, data_rows=20)
    _downgrade_meta_to_2_2_0(out)
    return out, str(mini_bundle.schema_path)


def _set_row(wb, sheet, row, **values):
    ws = wb[sheet]
    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for header, value in values.items():
        ws.cell(row, header_to_col[header]).value = value


def test_the_downgrade_helper_really_removes_the_new_keys(legacy_workbook):
    """Guard the guard: the fixture must actually produce an old-format file.

    If the downgrade silently stopped working, every test below would pass while
    testing nothing.
    """
    path, _ = legacy_workbook
    meta = _read_meta_rows(path)
    assert "node_order" not in meta
    assert "target_nodes" not in meta
    assert meta["path"] == "subject,visit,sample"


def test_legacy_workbook_validates_clean(legacy_workbook):
    """An empty old-format workbook still validates with no problems."""
    path, schema = legacy_workbook
    report = validate_workbook(path, schema)
    assert report.ok, [f.message for f in report.findings]


def test_legacy_workbook_reports_an_error_at_the_right_cell(legacy_workbook):
    """An old workbook's errors land on exactly the cell they always did.

    This is the real test of backwards compatibility: it proves the spec is
    rebuilt with the same columns in the same order, not merely that validation
    runs without crashing.
    """
    path, schema = legacy_workbook
    wb = openpyxl.load_workbook(path)
    _set_row(wb, "subject", 3, submitter_id="subj_1", subject_id="S1")
    _set_row(
        wb,
        "sample",
        3,
        submitter_id="samp_1",
        **{"subject.submitter_id": "ghost"},
        sample_id="X1",
        sample_type="Blood",
    )
    wb.save(path)

    report = validate_workbook(path, schema)
    dangling = [f for f in report.findings if f.validator == "link"]
    assert len(dangling) == 1
    assert dangling[0].sheet == "sample"
    assert dangling[0].cell.a1 == "B3"


def test_legacy_and_current_formats_produce_the_same_layout(mini_bundle, tmp_path):
    """Old and new metadata must rebuild an identical set of sheets and columns.

    Two copies of the same workbook — one with modern metadata, one downgraded —
    must validate to the same result. Any drift in column derivation between the
    two recovery paths would show up here.
    """
    spec = build_template_spec(mini_bundle, "sample", ["subject", "visit", "sample"])
    modern = tmp_path / "modern.xlsx"
    legacy = tmp_path / "legacy.xlsx"
    write_template(spec, modern, data_rows=20)
    write_template(spec, legacy, data_rows=20)
    _downgrade_meta_to_2_2_0(legacy)

    schema = str(mini_bundle.schema_path)
    modern_report = validate_workbook(modern, schema)
    legacy_report = validate_workbook(legacy, schema)

    assert modern_report.ok == legacy_report.ok
    assert sorted(modern_report.node_counts) == sorted(legacy_report.node_counts)


def test_node_order_wins_over_a_conflicting_legacy_path(mini_bundle, tmp_path):
    """When both formats are present the modern one decides.

    A workbook written by a newer g3mt carries both keys. ``node_order`` is the
    list the sheets were actually built from, so it must take precedence over
    the single-path key kept only for older readers.
    """
    spec = build_template_spec(mini_bundle, "sample", ["subject", "visit", "sample"])
    out = tmp_path / "both.xlsx"
    write_template(spec, out, data_rows=20)

    # Corrupt only the legacy key; node_order still describes the real sheets.
    wb = openpyxl.load_workbook(out)
    ws = wb[META_SHEET]
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "path":
            ws.cell(r, 2, "subject")
    wb.save(out)

    report = validate_workbook(out, str(mini_bundle.schema_path))
    assert set(report.node_counts) == {"subject", "visit", "sample"}


def test_workbook_with_no_meta_sheet_falls_back_to_path_arg(mini_bundle, tmp_path):
    """A workbook stripped of its metadata can still be checked with --path."""
    spec = build_template_spec(mini_bundle, "visit", ["subject", "visit"])
    out = tmp_path / "nometa.xlsx"
    write_template(spec, out, data_rows=20)

    wb = openpyxl.load_workbook(out)
    wb.remove(wb[META_SHEET])
    wb.save(out)

    report = validate_workbook(out, str(mini_bundle.schema_path), path_arg="subject,visit")
    assert report.ok


def test_workbook_with_no_meta_and_no_path_arg_raises(mini_bundle, tmp_path):
    """Without metadata or --path there is no way to know what the sheets are."""
    spec = build_template_spec(mini_bundle, "visit", ["subject", "visit"])
    out = tmp_path / "nometa.xlsx"
    write_template(spec, out, data_rows=20)

    wb = openpyxl.load_workbook(out)
    wb.remove(wb[META_SHEET])
    wb.save(out)

    with pytest.raises(WorkbookFormatError):
        validate_workbook(out, str(mini_bundle.schema_path))


def test_an_included_node_survives_validation(mini_bundle, tmp_path):
    """A sheet added with --include-node is still validated, not silently dropped.

    Because the recorded node list is trusted verbatim, a node the user
    deliberately brought back in (``project`` here, normally excluded) keeps its
    sheet at validation time. Re-applying the default exclusions would have
    thrown that sheet away.
    """
    spec = build_template_spec(
        mini_bundle,
        "subject",
        ["program", "project", "subject"],
        excluded_nodes=("program", "core_metadata_collection", "acknowledgement"),
    )
    out = tmp_path / "included.xlsx"
    write_template(spec, out, data_rows=20)

    report = validate_workbook(out, str(mini_bundle.schema_path))
    assert "project" in report.node_counts
