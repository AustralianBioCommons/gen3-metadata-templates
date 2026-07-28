"""End-to-end validation tests: generate a template, fill it, validate it.

This is the flagship test for the whole tool. It proves the pieces fit together:
a workbook produced by the writer can be read back, validated against the same
schema, and — crucially — that real mistakes are reported at the exact cell the
user needs to fix. A clean fill must pass; a deliberately broken fill must
produce precisely the expected findings and nothing spurious.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from gen3_metadata_templates import (
    build_multi_template_spec,
    build_template_spec,
    validate_workbook,
    write_template,
)
from gen3_metadata_templates.constants import DEFAULT_EXCLUDED_NODES
from gen3_metadata_templates.selection import resolve_selection
from gen3_metadata_templates.workbook.annotate import write_annotated_copy


def _set_row(wb, sheet, row, **values):
    ws = wb[sheet]
    header_to_col = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    for header, value in values.items():
        ws.cell(row, header_to_col[header]).value = value


@pytest.fixture()
def template_path(mini_bundle, tmp_path):
    """A freshly generated (empty) template for the subject->sample path."""
    spec = build_template_spec(mini_bundle, "sample", ["subject", "sample"])
    out = tmp_path / "roundtrip.xlsx"
    write_template(spec, out, data_rows=20)
    return out, str(mini_bundle.schema_path)


def test_valid_fill_passes(template_path):
    """A correctly filled workbook must validate with zero findings.

    If a clean submission produced spurious errors, users would lose trust in
    the tool immediately, so this is the most important guarantee.
    """
    path, schema = template_path
    wb = openpyxl.load_workbook(path)
    _set_row(wb, "subject", 3, submitter_id="subj_1", subject_id="S1", age=42, sex="Male")
    _set_row(
        wb,
        "sample",
        3,
        submitter_id="samp_1",
        **{"subject.submitter_id": "subj_1"},
        sample_id="X1",
        sample_type="Blood",
    )
    wb.save(path)

    report = validate_workbook(path, schema)
    assert report.ok
    assert report.findings == []


def test_broken_fill_reports_each_problem_at_the_right_cell(template_path):
    """Each planted error must surface as a finding located at its exact cell.

    The fill below plants five distinct mistakes; validation must find all five,
    each mapped to the cell and validator we expect — this is what lets a user
    fix "sheet subject, cell C4" rather than hunt through the file.
    """
    path, schema = template_path
    wb = openpyxl.load_workbook(path)
    # Row 3: valid baseline so the sample link has a real parent to match.
    _set_row(wb, "subject", 3, submitter_id="subj_1", subject_id="S1", age=30, sex="Female")
    # Row 4: duplicate submitter_id, non-integer age, invalid enum.
    _set_row(wb, "subject", 4, submitter_id="subj_1", subject_id="S2", age="ten", sex="Alien")
    # Sample row: dangling parent link and a missing required enum (sample_type).
    _set_row(
        wb, "sample", 3, submitter_id="samp_1", **{"subject.submitter_id": "ghost"}, sample_id="X1"
    )
    wb.save(path)

    report = validate_workbook(path, schema)
    assert not report.ok

    located = {(f.sheet, f.cell.a1 if f.cell else None): f.validator for f in report.findings}

    assert located.get(("subject", "C4")) == "type"  # age "ten"
    assert located.get(("subject", "G4")) == "enum"  # sex "Alien"
    assert located.get(("subject", "A4")) == "duplicate"  # repeated submitter_id
    assert located.get(("sample", "B3")) == "link"  # dangling subject link
    # Columns: A submitter_id, B subject.submitter_id, C sample_id, D sample_type.
    assert located.get(("sample", "D3")) == "required"  # empty sample_type


def test_annotated_copy_highlights_bad_cells(template_path, tmp_path):
    """The annotated workbook must fill each bad cell and attach a comment.

    Spreadsheet-native users fix errors fastest when they can open a copy and see
    the red cells, so the annotation must actually land on the flagged cells.
    """
    path, schema = template_path
    wb = openpyxl.load_workbook(path)
    _set_row(wb, "subject", 3, submitter_id="subj_1", subject_id="S1", age="oops", sex="Male")
    _set_row(
        wb,
        "sample",
        3,
        submitter_id="samp_1",
        **{"subject.submitter_id": "subj_1"},
        sample_id="X1",
        sample_type="Blood",
    )
    wb.save(path)

    report = validate_workbook(path, schema)
    annotated = tmp_path / "annotated.xlsx"
    write_annotated_copy(path, report, annotated)

    check = openpyxl.load_workbook(annotated)
    bad_cell = check["subject"]["C3"]
    assert bad_cell.fill.fgColor.rgb.endswith("FFC7CE")
    assert bad_cell.comment is not None
    assert "Validation Errors" in check.sheetnames


def test_annotate_refuses_to_overwrite_input(template_path):
    """Annotating over the input file would destroy the user's work — refuse it."""
    from gen3_metadata_templates.errors import G3mtError

    path, schema = template_path
    report = validate_workbook(path, schema)
    with pytest.raises(G3mtError):
        write_annotated_copy(path, report, path)


def test_schema_version_mismatch_warns_but_does_not_fail(template_path, tmp_path):
    """Validating against a different schema version warns, but is not an error.

    A template records the dictionary version it was built from. If someone later
    validates it against a bumped schema, they should be told (so they can
    regenerate) — but the file can still be structurally valid, so it's a warning,
    not a failure.
    """
    import json

    path, schema = template_path
    # Fill the template correctly so there are no real findings.
    wb = openpyxl.load_workbook(path)
    _set_row(wb, "subject", 3, submitter_id="subj_1", subject_id="S1", age=42, sex="Male")
    _set_row(
        wb,
        "sample",
        3,
        submitter_id="samp_1",
        **{"subject.submitter_id": "subj_1"},
        sample_id="X1",
        sample_type="Blood",
    )
    wb.save(path)

    # Build a copy of the schema with a different dictionary version.
    data = json.loads(Path(schema).read_text())
    data["_settings.yaml"]["_dict_version"] = "9.9.9"
    bumped = tmp_path / "bumped_schema.json"
    bumped.write_text(json.dumps(data))

    report = validate_workbook(path, str(bumped))
    assert report.ok  # structurally still valid
    assert any("9.9.9" in w and "0.1.0" in w for w in report.warnings)


# --- multi-node templates -------------------------------------------------


def _generate_clinical(bundle, tmp_path, name="clinical.xlsx"):
    """Generate a whole-category clinical template and return (path, schema)."""
    selection = resolve_selection(
        bundle,
        bundle.nodes_in_category("clinical"),
        excluded_nodes=DEFAULT_EXCLUDED_NODES,
        category="clinical",
    )
    spec = build_multi_template_spec(bundle, selection)
    out = tmp_path / name
    write_template(spec, out, data_rows=20)
    return out, str(bundle.schema_path)


def test_multi_target_clinical_roundtrip(clinical_hub_bundle, tmp_path):
    """The researcher's whole journey: one command, fill it in, validate clean.

    This mirrors real clinical data: one participant seen at two timepoints, with
    a measurement recorded at each. Two ``clinical_descriptor`` rows share the
    same ``subject.submitter_id`` — that reuse *is* the one-to-many relationship,
    and it must validate without complaint.
    """
    path, schema = _generate_clinical(clinical_hub_bundle, tmp_path)
    wb = openpyxl.load_workbook(path)
    _set_row(wb, "subject", 3, submitter_id="subj_1", patient_id="P01")
    _set_row(
        wb,
        "clinical_descriptor",
        3,
        submitter_id="cd_1",
        **{"subject.submitter_id": "subj_1"},
        timepoint_label="baseline",
    )
    _set_row(
        wb,
        "clinical_descriptor",
        4,
        submitter_id="cd_2",
        **{"subject.submitter_id": "subj_1"},
        timepoint_label="year_2",
    )
    _set_row(
        wb,
        "demographic",
        3,
        submitter_id="demo_1",
        **{"clinical_descriptor.submitter_id": "cd_1"},
        sex="Male",
    )
    _set_row(
        wb,
        "demographic",
        4,
        submitter_id="demo_2",
        **{"clinical_descriptor.submitter_id": "cd_2"},
        sex="Male",
    )
    wb.save(path)

    report = validate_workbook(path, schema)
    assert report.ok, [f"{f.location}: {f.message}" for f in report.findings]


def test_multi_target_roundtrip_reports_a_dangling_link(clinical_hub_bundle, tmp_path):
    """A measurement pointing at a timepoint that doesn't exist is caught.

    Across many sheets it's easy to mistype a parent id. The error must name the
    exact cell so the fix is obvious.
    """
    path, schema = _generate_clinical(clinical_hub_bundle, tmp_path)
    wb = openpyxl.load_workbook(path)
    _set_row(wb, "subject", 3, submitter_id="subj_1", patient_id="P01")
    _set_row(
        wb,
        "clinical_descriptor",
        3,
        submitter_id="cd_1",
        **{"subject.submitter_id": "subj_1"},
        timepoint_label="baseline",
    )
    _set_row(
        wb,
        "demographic",
        3,
        submitter_id="demo_1",
        **{"clinical_descriptor.submitter_id": "ghost"},
        sex="Male",
    )
    wb.save(path)

    report = validate_workbook(path, schema)
    dangling = [f for f in report.findings if f.validator == "link"]
    assert len(dangling) == 1
    assert dangling[0].sheet == "demographic"
    assert dangling[0].cell.a1 == "B3"


def test_an_unfilled_sheet_in_a_multi_target_template_is_not_an_error(
    clinical_hub_bundle, tmp_path
):
    """Sheets you have no data for can be left empty.

    Asking for a whole category gives sheets a given study may not use. The
    Instructions sheet promises those can be left blank, so validation must
    honour that promise.
    """
    path, schema = _generate_clinical(clinical_hub_bundle, tmp_path)
    wb = openpyxl.load_workbook(path)
    _set_row(wb, "subject", 3, submitter_id="subj_1", patient_id="P01")
    _set_row(
        wb,
        "clinical_descriptor",
        3,
        submitter_id="cd_1",
        **{"subject.submitter_id": "subj_1"},
        timepoint_label="baseline",
    )
    wb.save(path)  # demographic, blood_pressure_test, medical_history all left empty

    report = validate_workbook(path, schema)
    assert report.ok, [f"{f.location}: {f.message}" for f in report.findings]


def test_branching_multi_target_roundtrip(mini_bundle, tmp_path):
    """A node with two parents in the template gets a column for each.

    Selecting both ``visit`` and ``assay_file`` pulls ``sample`` in with two
    possible parents. Only the one the schema requires is mandatory, so filling
    just that one must validate cleanly.
    """
    selection = resolve_selection(
        mini_bundle, ["visit", "assay_file"], excluded_nodes=DEFAULT_EXCLUDED_NODES
    )
    spec = build_multi_template_spec(mini_bundle, selection)
    out = tmp_path / "branching.xlsx"
    write_template(spec, out, data_rows=20)

    sample = spec.node_template("sample")
    assert sample.column_by_header("subject.submitter_id").required is True
    assert sample.column_by_header("visit.submitter_id").required is False

    wb = openpyxl.load_workbook(out)
    _set_row(wb, "subject", 3, submitter_id="subj_1", subject_id="S1")
    _set_row(
        wb, "visit", 3, submitter_id="v_1", **{"subject.submitter_id": "subj_1"}, visit_id="V1"
    )
    _set_row(
        wb,
        "sample",
        3,
        submitter_id="samp_1",
        **{"subject.submitter_id": "subj_1"},
        sample_id="X1",
        sample_type="Blood",
    )
    _set_row(
        wb,
        "assay_file",
        3,
        submitter_id="af_1",
        **{"sample.submitter_id": "samp_1"},
        file_name="reads.bam",
    )
    wb.save(out)

    report = validate_workbook(out, str(mini_bundle.schema_path))
    assert report.ok, [f"{f.location}: {f.message}" for f in report.findings]
