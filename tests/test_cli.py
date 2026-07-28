"""Tests for the ``g3mt`` command-line interface.

These drive the CLI the way a user would (via Typer's test runner) to confirm
the command surface behaves: templates get written, ambiguity is refused with a
helpful message in a non-interactive run, and the exit codes that scripts rely
on (0 clean, 1 problems, 2 usage error) are correct.
"""

from __future__ import annotations

import json

import openpyxl
from typer.testing import CliRunner

from gen3_metadata_templates.cli import app

runner = CliRunner()


def test_generate_writes_a_workbook(mini_schema_path, tmp_path):
    """`g3mt generate` with an unambiguous path writes the .xlsx and exits 0."""
    out = tmp_path / "visit.xlsx"
    result = runner.invoke(app, ["generate", mini_schema_path, "visit", "-o", str(out)])
    assert result.exit_code == 0
    assert out.exists()
    assert "visit" in openpyxl.load_workbook(out).sheetnames


def test_generate_ambiguous_without_path_exits_2(mini_schema_path, tmp_path):
    """An ambiguous target with no --path fails clearly in a non-interactive run.

    The runner provides no TTY, so the CLI must not hang on a prompt; it should
    print the numbered options and exit 2 so a script knows to pass --path.
    """
    out = tmp_path / "sample.xlsx"
    result = runner.invoke(app, ["generate", mini_schema_path, "sample", "-o", str(out)])
    assert result.exit_code == 2
    assert "multiple paths" in result.output.lower()


def test_generate_with_path_index(mini_schema_path, tmp_path):
    """Passing --path resolves the ambiguity and writes the chosen sheets."""
    out = tmp_path / "sample.xlsx"
    result = runner.invoke(
        app, ["generate", mini_schema_path, "sample", "--path", "2", "-o", str(out)]
    )
    assert result.exit_code == 0
    names = openpyxl.load_workbook(out).sheetnames
    assert "visit" in names  # path 2 is subject -> visit -> sample


def test_list_paths_prints_and_exits_zero(mini_schema_path):
    """`--list-paths` shows the numbered options without generating anything."""
    result = runner.invoke(app, ["generate", mini_schema_path, "sample", "--list-paths"])
    assert result.exit_code == 0
    assert "1." in result.output and "2." in result.output


def test_validate_clean_workbook_exits_zero(mini_schema_path, tmp_path):
    """A generated (empty) template has no rows, so validation is clean (exit 0)."""
    out = tmp_path / "empty.xlsx"
    runner.invoke(app, ["generate", mini_schema_path, "visit", "-o", str(out)])
    result = runner.invoke(app, ["validate", str(out), "-s", mini_schema_path])
    assert result.exit_code == 0
    assert "all good" in result.output.lower()


def test_validate_reports_problems_exits_one(mini_schema_path, tmp_path):
    """A workbook with a bad value exits 1 and names the problem.

    Exit 1 (distinct from usage errors) is how CI or a script can tell "the file
    has fixable problems" apart from "you called me wrong".
    """
    out = tmp_path / "bad.xlsx"
    runner.invoke(app, ["generate", mini_schema_path, "visit", "-o", str(out)])
    wb = openpyxl.load_workbook(out)
    ws = wb["visit"]
    header = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    ws.cell(3, header["submitter_id"]).value = "v1"
    ws.cell(3, header["visit_id"]).value = "V1"
    ws.cell(3, header["subject.submitter_id"]).value = "ghost"  # dangling link
    wb.save(out)

    result = runner.invoke(app, ["validate", str(out), "-s", mini_schema_path])
    assert result.exit_code == 1
    assert "problem" in result.output.lower()


def test_validate_json_output_is_parseable(mini_schema_path, tmp_path):
    """`--json` emits machine-readable output for scripting / a future UI."""
    out = tmp_path / "empty.xlsx"
    runner.invoke(app, ["generate", mini_schema_path, "visit", "-o", str(out)])
    result = runner.invoke(app, ["validate", str(out), "-s", mini_schema_path, "--json"])
    payload = json.loads(result.output)
    assert payload["ok"] is True


def test_error_without_debug_is_clean(tmp_path):
    """By default an input error is a one-line message, not a traceback.

    Non-developers shouldn't be shown a Python stack trace; they get a plain
    message, exit code 2, and a hint that --debug exists.
    """
    missing = str(tmp_path / "nope.json")
    result = runner.invoke(app, ["nodes", missing])
    assert result.exit_code == 2
    assert "Error:" in result.output
    assert "Traceback" not in result.output
    assert "--debug" in result.output


def test_debug_flag_shows_traceback(tmp_path):
    """`--debug` swaps the clean message for a full traceback (still exit 2).

    When something goes wrong and the plain message isn't enough, --debug gives
    a developer the stack trace to diagnose it, without changing the exit code.
    """
    missing = str(tmp_path / "nope.json")
    result = runner.invoke(app, ["--debug", "nodes", missing])
    assert result.exit_code == 2
    assert "Traceback" in result.output


def test_categories_command_lists_categories_with_counts(mini_schema_path):
    """`g3mt categories` shows what groups of nodes the schema defines.

    This is the discovery step for someone who knows they want "clinical data"
    but not which node names that involves.
    """
    result = runner.invoke(app, ["categories", mini_schema_path])
    assert result.exit_code == 0
    assert "clinical" in result.output
    assert "administrative" in result.output


def test_categories_command_can_hide_node_names(mini_schema_path):
    """`--no-nodes` gives just the counts, for a schema with many nodes."""
    result = runner.invoke(app, ["categories", mini_schema_path, "--no-nodes"])
    assert result.exit_code == 0
    assert "clinical" in result.output
    assert "assay_file" not in result.output


def test_nodes_command_shows_the_category_column(mini_schema_path):
    """`g3mt nodes` surfaces each node's category alongside its links."""
    result = runner.invoke(app, ["nodes", mini_schema_path])
    assert result.exit_code == 0
    assert "Category" in result.output
    assert "biospecimen" in result.output


# --- multi-node selection -------------------------------------------------
#
# The existing single-target tests above are the backwards-compatibility gate:
# `g3mt generate schema.json sample` must keep prompting/refusing exactly as it
# did. These cover the new selection flags.


def test_generate_with_category_writes_every_node_in_it(mini_schema_path, tmp_path):
    """`--category` builds one workbook covering the whole category.

    This is the headline case: a researcher submitting clinical data runs one
    command instead of hunting for node names and generating several files.
    """
    out = tmp_path / "clinical.xlsx"
    result = runner.invoke(
        app, ["generate", mini_schema_path, "--category", "clinical", "-o", str(out)]
    )
    assert result.exit_code == 0
    names = openpyxl.load_workbook(out).sheetnames
    assert "visit" in names  # the clinical node
    assert "subject" in names  # its ancestor, pulled in automatically


def test_generate_with_repeated_node_flags_unions_them(mini_schema_path, tmp_path):
    """Several --node flags produce one workbook with the union of their paths."""
    out = tmp_path / "union.xlsx"
    result = runner.invoke(
        app,
        ["generate", mini_schema_path, "--node", "visit", "--node", "sample", "-o", str(out)],
    )
    assert result.exit_code == 0
    names = openpyxl.load_workbook(out).sheetnames
    assert {"subject", "visit", "sample"}.issubset(set(names))


def test_multi_target_never_prompts_or_refuses_on_ambiguity(mini_schema_path, tmp_path):
    """A multi-node selection resolves ambiguity itself instead of exiting 2.

    ``sample`` has two possible paths. In single-target mode that's an error
    without --path; in a multi-node selection the shortest is taken and the
    alternative is reported, so a whole-category run can't be derailed.
    """
    out = tmp_path / "amb.xlsx"
    result = runner.invoke(app, ["generate", mini_schema_path, "--node", "sample", "-o", str(out)])
    assert result.exit_code == 0
    assert "--path sample=" in result.output


def test_generate_without_any_selection_exits_2(mini_schema_path):
    """Naming no node at all is a usage error that shows how to fix it."""
    result = runner.invoke(app, ["generate", mini_schema_path])
    assert result.exit_code == 2
    assert "--category" in result.output


def test_generate_unknown_category_lists_what_is_available(mini_schema_path):
    """A wrong category name shows the real ones rather than failing blankly."""
    result = runner.invoke(app, ["generate", mini_schema_path, "--category", "nonsense"])
    assert result.exit_code == 2
    assert "clinical" in result.output


def test_generate_explicit_node_that_is_also_excluded_exits_2(mini_schema_path, tmp_path):
    """Asking for a node and excluding it in the same command is contradictory."""
    result = runner.invoke(
        app,
        [
            "generate",
            mini_schema_path,
            "--node",
            "sample",
            "--exclude-node",
            "sample",
            "-o",
            str(tmp_path / "x.xlsx"),
        ],
    )
    assert result.exit_code == 2


def test_category_member_excluded_is_skipped_with_a_note(mini_schema_path, tmp_path):
    """Excluding one member of a category trims it and says so."""
    out = tmp_path / "trimmed.xlsx"
    result = runner.invoke(
        app,
        [
            "generate",
            mini_schema_path,
            "--category",
            "clinical",
            "--exclude-node",
            "visit",
            "-o",
            str(out),
        ],
    )
    # Every clinical node was excluded, so there is nothing left to generate.
    assert result.exit_code == 2
    assert "clinical" in result.output


def test_path_override_for_one_target_of_several(mini_schema_path, tmp_path):
    """`--path node=N` picks a route for one target without affecting the others."""
    out = tmp_path / "override.xlsx"
    result = runner.invoke(
        app,
        [
            "generate",
            mini_schema_path,
            "--node",
            "sample",
            "--path",
            "sample=2",
            "-o",
            str(out),
        ],
    )
    assert result.exit_code == 0
    assert "visit" in openpyxl.load_workbook(out).sheetnames


def test_bare_path_with_multiple_targets_exits_2(mini_schema_path, tmp_path):
    """A bare --path is ambiguous when several nodes were selected."""
    result = runner.invoke(
        app,
        [
            "generate",
            mini_schema_path,
            "--node",
            "sample",
            "--node",
            "visit",
            "--path",
            "2",
            "-o",
            str(tmp_path / "x.xlsx"),
        ],
    )
    assert result.exit_code == 2
    assert "--path" in result.output


def test_default_output_name_for_a_category(mini_schema_path, tmp_path, monkeypatch):
    """A category selection names the file after the category."""
    monkeypatch.chdir(tmp_path)
    result = runner.invoke(app, ["generate", mini_schema_path, "--category", "clinical"])
    assert result.exit_code == 0
    assert (tmp_path / "clinical_template.xlsx").exists()


def test_default_output_name_for_the_positional_form_is_unchanged(
    mini_schema_path, tmp_path, monkeypatch
):
    """The original single-target filename convention still applies."""
    monkeypatch.chdir(tmp_path)
    result = runner.invoke(app, ["generate", mini_schema_path, "visit"])
    assert result.exit_code == 0
    assert (tmp_path / "visit_template.xlsx").exists()


def test_list_paths_with_multiple_targets_groups_by_node(mini_schema_path):
    """`--list-paths` groups the options under each selected node."""
    result = runner.invoke(
        app,
        ["generate", mini_schema_path, "--node", "sample", "--node", "visit", "--list-paths"],
    )
    assert result.exit_code == 0
    assert "sample" in result.output and "visit" in result.output


def test_generate_category_then_validate_exits_zero(mini_schema_path, tmp_path):
    """The whole journey works: select a category, generate, validate clean."""
    out = tmp_path / "journey.xlsx"
    generated = runner.invoke(
        app, ["generate", mini_schema_path, "--category", "clinical", "-o", str(out)]
    )
    assert generated.exit_code == 0

    validated = runner.invoke(app, ["validate", str(out), "-s", mini_schema_path])
    assert validated.exit_code == 0
